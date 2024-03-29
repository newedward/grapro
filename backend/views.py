from django.shortcuts import render
import json
from django.http import HttpResponse
from django.http import JsonResponse
from django.core import serializers
from django.http import StreamingHttpResponse
from django.utils.encoding import escape_uri_path
from backend.models import *
from backend.controller import dbcontrol
from backend.controller import Util
import os
import time
import openpyxl
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# Create your views here.

def getCurUserID(req):
    response = {}
    err, curUser = Util.getUserIDBySession(req)
    info,role = dbcontrol.getUserRole(curUser)
    info2,avater = dbcontrol.getAvaterByID(curUser)
    if err == 'succeed':
        response['userID'] = curUser
        response['msg'] = 'succeed'
        response['role'] = role
        response['err_num'] = 0
        response['avater'] = avater
    else:
        response['err_num'] = 1
        response['msg'] = 'error'
    # print(response)
    return JsonResponse(response)

def getCurUserIDMan(req):
    response = {}
    err, curUser = Util.getUserIDBySession(req)
    info,role = dbcontrol.getUserRole(curUser)
    now,recommend = dbcontrol.getcountbyId(curUser)
    if err == 'succeed':
        response['userID'] = curUser
        response['msg'] = 'succeed'
        response['role'] = role
        response['err_num'] = 0
        response['avater'] = now
        response['count'] = recommend
    else:
        response['err_num'] = 1
        response['msg'] = 'error'
    # print(response)
    return JsonResponse(response)

def getUserAvaterByID(req):
    userid = req.POST.get('watchId')
    info,avater = dbcontrol.getAvaterByID(userid)
    response = {}
    if info == "succeed":
        response['Msg'] = 'succeed'
        response['err_code'] = 0
        response['avater'] = avater
    else :
        response['Msg'] = 'failed'
        response['err_code'] = 1
    # print("到这",response)
    return JsonResponse(response)

def getTeacher(req):
    response = {}
    user = str(req.POST.get('userId'))
    # 根据user得到他院系所有的老师
    info,tealist,ulist= dbcontrol.getTeaByStu(user)
    if info == "succeed":
        response['Msg'] = 'succeed'
        response['err_code'] = 0
        response['tealist'] = json.loads(serializers.serialize("json", tealist))
        response['ulist'] = json.loads(serializers.serialize("json", ulist))
    else :
        response['Msg'] = 'failed'
        response['err_code'] = 1
    # print(response)
    return JsonResponse(response)

def applyTeacher(req):
    teacher = req.POST.get('teacherId')
    student = req.POST.get('studentId')
    info = dbcontrol.addapplication(teacher,student)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    elif info == "has":
        response['Msg'] = 'has'
        response['err_code'] = 2
    elif info == "not start":
        response['Msg'] = 'not start'
        response['err_code'] = 3
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)

def getApplicationByTeacher(req):
    teacher = req.POST.get('teacherId')
    info,apps,stus,users = dbcontrol.getapplicationbytea(teacher)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
        response["apps"] = json.loads(serializers.serialize("json", apps))
        response["slist"] = json.loads(serializers.serialize("json", stus))
        response["ulist"] = json.loads(serializers.serialize("json", users))
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    # print(response)
    return JsonResponse(response)

def acceptStu(req):
    teacher = req.POST.get('teacherId')
    student = req.POST.get('stuId')
    app = req.POST.get('appId')
    # print(student,teacher,app)
    dbcontrol.checkapplication(app)
    info = dbcontrol.followteacher(student, teacher)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    elif info == "has":
        response['Msg'] = 'has'
        response['err_code'] = 2
    elif info == "wrong":
        response['Msg'] = 'wrong'
        response['err_code'] = 3
    else :
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)

def refuseStu(req):
    app = req.POST.get('appId')
    info = dbcontrol.checkapplication(app)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)

def addTeaQueue(req):
    teacher = req.POST.get('teacherId')
    student = req.POST.get('stuId')
    app = req.POST.get('appId')
    dbcontrol.checkapplication(app)
    info = dbcontrol.addteaqueue(student,teacher)
    print("add",info)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    elif info == "has":
        response['Msg'] = 'has'
        response['err_code'] = 2
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)

def getTeaQueue(req):
    teacher = req.POST.get('teacherId')
    info,queuestr,ulist,slist = dbcontrol.getteaqueue(teacher)
    response = {}
    if info == "succeed":
        queue = queuestr.split(',')
        response["queue"] = queue
        response["ulist"] = json.loads(serializers.serialize("json", ulist))
        response["slist"] = json.loads(serializers.serialize("json", slist))
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)

def storeTeaQueue(req):
    queue = req.POST.get('queue')
    teacher = req.POST.get('teaid')
    str = queue[1:-1]
    str += ","
    print(str)
    info = dbcontrol.storeteaqueue(teacher,str)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)

def addWorkTitle(req):
    title = req.POST.get('title')
    student = req.POST.get('stuId')
    info,workid = dbcontrol.addworkbytitle(title=title,stuid=student)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
        response["workid"] = workid
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)

def getRecordTitle(req):
    student = req.POST.get("stuId")
    process = req.POST.get("process")
    info,stu = dbcontrol.getStuById(student)
    info,relist =dbcontrol.getrecordbystu(student,process)
    response = {}
    if stu.work != None and info == "succeed":
        response['rlist'] = json.loads(serializers.serialize("json", relist))
        response['title'] = stu.work.title
        response['workid'] = stu.work.id
        response["Msg"] = "succeed"
        response["err_code"] = 0
    elif stu.work == None and info == "succeed":
        response['rlist'] = json.loads(serializers.serialize("json", relist))
        response['title'] = "null"
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)

def addRecordStu(req):
    stuid = req.POST.get("stuId")
    workid = req.POST.get("workId")
    process = req.POST.get("process")
    path = req.POST.get("path")
    info,recordid = dbcontrol.addrecordstu(stuid,workid,path,process)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
        response["recordid"] = recordid
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)

def addRecordIntroduction(req):
    recordid  = req.POST.get('recordId')
    introduction = req.POST.get('introduction')
    info= dbcontrol.addrecordintroduction(recordid, introduction)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    # print(response)
    return JsonResponse(response)

def addRecordContent(req):
    recordid = req.POST.get('recordId')
    content = req.POST.get('content')
    info = dbcontrol.addrecordcontent(recordid, content)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    print(response)
    return JsonResponse(response)

def downloadfile(req,id):
    print("here",id)
    def down_chunk_file_manager(file_path, chuck_size=1024):
        with open(file_path, "rb") as file:
            while True:
                chuck_stream = file.read(chuck_size)
                if chuck_stream:
                    yield chuck_stream
                else:
                    break

    file_path = dbcontrol.getpathbyid(id)
    if not os.path.exists(file_path):
        return "no file 联系开发者"
    names = file_path.split('/')
    name = names[-1]
    print(name)
    response = StreamingHttpResponse(down_chunk_file_manager(file_path))
    response['Content-Type'] = 'application/octet-stream'
    response["Content-Disposition"] = "attachment; filename*=UTF-8''{}".format(escape_uri_path(name))

    return response

def uploadfile(req):
    work = req.FILES.get('work',None)
    nextnum = req.POST.get('next')
    username = req.POST.get('stuid')
    process = req.POST.get('process')
    print("下一个文件",nextnum)
    # head_path = BASE_DIR + "\\{}".format(process).replace(" ", "")
    head_path = BASE_DIR + "/media/{}/{}/{}".format(process,username,nextnum).replace(" ", "")
    print("head_path", head_path)
    if not os.path.exists(head_path):
        os.makedirs(head_path)
    suffix = work.name.split(".")[1]
    work_path = head_path + "/"+work.name
    work_path = work_path.replace(" ", "")
    print("路径",work_path)
    response = {}
    try:
        with open(work_path, 'wb') as f:
            for chunk in work.chunks():
                f.write(chunk)
    except:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    else:
        response["Msg"] = "succeed"
        response["err_code"] = 0
        response["path"] = work_path
    return JsonResponse(response)

        # # 图片后缀
        # head_suffix = file_obj.name.split(".")[1]
        # print("图片后缀", head_suffix)  # 图片后缀 png
        #
        # # 储存路径
        # file_path = head_path + "\\{}".format("head." + head_suffix)
        # file_path = file_path.replace(" ", "")
        # print("储存路径", file_path)  # C:\Users\user\Desktop\DownTest\media\username\head\head.png
def getMyStudent(req):
    teaid = req.POST.get("teacherId")
    info,slist,ulist = dbcontrol.getStuByTea(teaid)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
        response["slist"] = json.loads(serializers.serialize("json", slist))
        response["ulist"] = json.loads(serializers.serialize("json", ulist))
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    # print("my",response)
    return JsonResponse(response)

def getStudentEndbyTea(req):
    teaid = req.POST.get("teaId")
    info, slist, ulist = dbcontrol.getStuByTea(teaid)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
        response["ulist"] = json.loads(serializers.serialize("json", ulist))
        response['slist'] = json.loads(serializers.serialize("json", slist))
        reclist = []
        ulist2 = []
        slist2 = []
        for s in slist:
            info, rec = dbcontrol.getlatstrecordbystu(s,'E')
            if info == "succeed":
                reclist.append(rec)
        for rec in reclist:
            ulist2.append(User.objects.get(id=rec.student_id))
            slist2.append(rec.student)
        response["rlist"] = json.loads(serializers.serialize("json", reclist))
        response["ulist"] = json.loads(serializers.serialize("json", ulist2))
        response['slist'] = json.loads(serializers.serialize("json", slist2))
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    # print(response)
    return JsonResponse(response)

def getStudentMediumbyTea(req):
    teaid = req.POST.get("teaId")
    info, slist, ulist = dbcontrol.getStuByTea(teaid)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
        response["ulist"] = json.loads(serializers.serialize("json", ulist))
        response['slist'] = json.loads(serializers.serialize("json", slist))
        reclist = []
        ulist2 = []
        slist2 = []
        for s in slist:
            info, rec = dbcontrol.getlatstrecordbystu(s,'M')
            if info == "succeed":
                reclist.append(rec)
        for rec in reclist:
            ulist2.append(User.objects.get(id=rec.student_id))
            slist2.append(rec.student)
        response["rlist"] =json.loads(serializers.serialize("json", reclist))
        response["ulist"] = json.loads(serializers.serialize("json", ulist2))
        response['slist'] = json.loads(serializers.serialize("json", slist2))
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    # print(response)
    return JsonResponse(response)

def getStudentStartbyTea(req):
    teaid = req.POST.get("teaId")
    info,slist,ulist = dbcontrol.getStuByTea(teaid)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
        reclist = []
        ulist2 = []
        slist2 = []
        for s in slist:
            info,rec = dbcontrol.getlatstrecordbystu(s,'S')
            if info == "succeed":
                reclist.append(rec)
        for rec in reclist:
            ulist2.append(User.objects.get(id=rec.student_id))
            slist2.append(rec.student)
        response["rlist"] =json.loads(serializers.serialize("json", reclist))
        response["ulist"] = json.loads(serializers.serialize("json", ulist2))
        response['slist'] = json.loads(serializers.serialize("json", slist2))
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    # print(response)
    return JsonResponse(response)

def login(req):
    userName = str( req.POST.get('userName') )
    password = str( req.POST.get('password') )

    err, curUser = dbcontrol.login(userName, password)
    print(err)
    if (err!="succeed"):
        return HttpResponse(err)

    Util.setUserForSession(req, curUser.id)

    return HttpResponse(err)

def logout(req):
    info = Util.delUserForSession(req)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)

def validateUserName(req):
    username = req.POST.get('username')
    info = dbcontrol.validateUser(username)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)

def createUser(req):
    userName = str(req.POST.get('userName'))
    password = str(req.POST.get('password'))
    name = str(req.POST.get('name'))
    email = str(req.POST.get('email'))
    avatarpath = str(req.POST.get('avatarpath'))
    uni = str(req.POST.get('uni'))
    school = str(req.POST.get('school'))
    role = str(req.POST.get('role'))
    requirement = str(req.POST.get('req'))
    code = str(req.POST.get('code'))
    avatarpath = dbcontrol.changeavater(avatarpath)
    info,user = dbcontrol.addUser(userName,password,name,email,avatarpath,uni,school)
    if info == "succeed":
        if role == "学生":
            user.credit = 1
            user.save()
            info = dbcontrol.addStudent(user,code)
            if info == "succeed":
                # print(Util.setUserForSession(req, user.id))
                return HttpResponse("注册成功!但仍需要管理员审批")
            else:
                print("学生失败")
                return HttpResponse("注册失败!")
        else :
            user.credit = 2
            user.save()
            info = dbcontrol.addTeacher(user,requirement)
            if info == "succeed":
                # print(Util.setUserForSession(req, user.id))
                return HttpResponse("注册成功!但仍需要管理员审批")
            else:
                print("教师失败")
                return HttpResponse("注册失败!")
    else:
        print("用户失败")
        return HttpResponse("注册失败!")


def uploadAvater(req):
    avater = req.FILES.get('avater', None)
    timestamp = time.time()
    head_path = BASE_DIR + "/allstatic/{}".format(timestamp).replace(" ", "")
    read_path = BASE_DIR + "/static/{}".format(timestamp).replace(" ", "")
    if not os.path.exists(head_path):
        os.makedirs(head_path)
    suffix = avater.name.split(".")[1]
    work_path = head_path + "/" + avater.name
    read_path2 = read_path + "/" + avater.name
    ava_path = work_path.replace(" ", "")
    read_path3 = read_path2.replace(" ", "")
    print("路径0", ava_path)
    response = {}
    try:
        with open(ava_path, 'wb') as f:
            for chunk in avater.chunks():
                f.write(chunk)
    except:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    else:
        response["Msg"] = "succeed"
        response["err_code"] = 0
        response["path"] = read_path3
    return JsonResponse(response)

def changeInfo(req):
    name = str(req.POST.get('name'))
    email = str(req.POST.get('email'))
    avatarpath = str(req.POST.get('avatarpath'))
    requirement = str(req.POST.get('req'))
    code = str(req.POST.get('code'))
    id = str(req.POST.get('userId'))
    role =str(req.POST.get('roleId'))
    avatarpath = dbcontrol.changeavater(avatarpath)
    print("信息",id,avatarpath,requirement)
    info = dbcontrol.changeuserinfo(id,role,avatarpath,requirement,email,name,code)
    if info == "succeed":
        return HttpResponse("修改成功!")
    else:
        return HttpResponse("修改失败!")


def getUser(req):
    userid = req.POST.get('watchId')
    info,user,extra = dbcontrol.getUser(userid)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_num"] = 0
        response["name"] = user.name
        response["username"] = user.username
        response["uni"] = user.uni
        response["school"] = user.school
        response["email"] = user.email
        response["req"] = extra
        response["code"] = extra
    else:
        response['Msg'] = 'failed'
        response['err_num'] = 1
    return JsonResponse(response)

def getTeacherProcess(req):
    teaid = req.POST.get('userId')
    info, time1,time2,time3,stalist = dbcontrol.getteaprocess(teaid)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
        response["time1"] = time1
        response["time2"] = time2
        response["time3"] = time3
        response["stalist"] = stalist
    elif info == "no":
        response['Msg'] = 'no stu'
        response['err_code'] = 1
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 2
    print(response)
    return JsonResponse(response)

def subTeacherProcess(req):
    teaid = req.POST.get('userId')
    time1 = req.POST.get('time1')
    time2 = req.POST.get('time2')
    time3 = req.POST.get('time3')
    info = dbcontrol.subteaprocess(teaid,time1,time2,time3)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)

def getStudentStartProcess(req):
    stuid = req.POST.get('stuId')
    stu = Student.objects.get(user_stu_id=stuid)
    info,type = Util.getStatusItemStuStart(stu.status, stu.work)
    response = {}
    response["info"] = info
    response["type"] = type
    response['statuscode'] = stu.status
    return JsonResponse(response)

def getStudentMediumProcess(req):
    stuid = req.POST.get('stuId')
    stu = Student.objects.get(user_stu_id=stuid)
    info, type = Util.getStatusItemStuMedium(stu.status, stu.work)
    response = {}
    response["info"] = info
    response["type"] = type
    response['statuscode'] = stu.status
    return JsonResponse(response)

def getStudentEndProcess(req):
    stuid = req.POST.get('stuId')
    stu = Student.objects.get(user_stu_id=stuid)
    info, type = Util.getStatusItemStuEnd(stu.status, stu.work)
    response = {}
    response["info"] = info
    response["type"] = type
    response['statuscode'] = stu.status
    return JsonResponse(response)

def initFileMan(req):
    manid = req.POST.get('manId')
    info,data = dbcontrol.showarchive(manid)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
        response["data"] = data
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    # print(response)
    return JsonResponse(response)

def initHome(req):
    watchid = req.POST.get('watchId')
    info, data = dbcontrol.inithome(watchid)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
        response["data"] = data
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    # print(response)
    return JsonResponse(response)

def initcheckman(req):
    manid = req.POST.get('manId')
    info,clist,nlist,ids = dbcontrol.initcheckman(manid)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
        response["nlist"] = nlist
        response["clist"] = clist
        response['ids'] = ids
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    print(response)
    return JsonResponse(response)

def checkUser(req):
    ids = str(req.POST.get('ids'))
    idlist = filter(None,ids.split(','))
    info = dbcontrol.checkuser(idlist)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    print(response)
    return JsonResponse(response)

def delUser(req):
    ids = str(req.POST.get('ids'))
    idlist = filter(None, ids.split(','))
    info = dbcontrol.deluser(idlist)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)

def storeArchive(req):
    recid = req.POST.get('recId')
    content = req.POST.get('content')
    teaid = req.POST.get('teaId')
    info = dbcontrol.addarchive(recid,content,teaid)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    print(response)
    return JsonResponse(response)

def loginBacth(req):
    sheet = req.FILES.get('loginsheet', None)
    manid = req.POST.get('manid')
    suffix = req.POST.get('suffix')
    response = {}
    wb = openpyxl.load_workbook(sheet)
    ws = wb.active
    if sheet.name == "学生.xlsx" or sheet.name == "学生.xls":
        code = ws['A']
        name = ws['B']
        for c,n in zip(code,name):
            info = dbcontrol.registerStu(n.value,c.value,manid,suffix)
            if info != "succeed":
                response['Msg'] = 'error'
                response['err_code'] = 1
                return JsonResponse(response)
        response["Msg"] = "succeed"
        response["err_code"] = 0
        return JsonResponse(response)
    elif sheet.name == "导师.xlsx" or sheet.name == "导师.xls":
        name = ws['A']
        username = ws['B']
        for n,u in zip(name,username):
            info = dbcontrol.registerTea(u.value,suffix,n.value,manid)
            if info != "succeed":
                response['Msg'] = 'error'
                response['err_code'] = 1
                return JsonResponse(response)
        response["Msg"] = "succeed"
        response["err_code"] = 0
        return JsonResponse(response)
    else:
        response['Msg'] = 'name error'
        response['err_code'] = -1
    return JsonResponse(response)


def makeGreatPro(req):
    stuid = req.POST.get('stuId')
    info = dbcontrol.makegreatproject(stuid)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    # print(response)
    return JsonResponse(response)

def startnow(req):
    manid = req.POST.get('manId')
    count = req.POST.get('count')
    info = dbcontrol.startnowbyman(manid,count)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)


def endnow(req):
    manid = req.POST.get('manId')
    info = dbcontrol.endnowbyman(manid)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)

def startprocess(req):
    manid = req.POST.get('manId')
    info = dbcontrol.startprocess(manid)
    response = {}
    if info == "succeed":
        response["Msg"] = "succeed"
        response["err_code"] = 0
    else:
        response['Msg'] = 'failed'
        response['err_code'] = 1
    return JsonResponse(response)







